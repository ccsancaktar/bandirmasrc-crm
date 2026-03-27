from fastapi import FastAPI, Depends, HTTPException, status, Query, Body
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel
from datetime import datetime, timedelta
from typing import Optional, List, Union, Any
import jwt
from passlib.context import CryptContext
import sqlite3
from dotenv import load_dotenv
import os
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from fastapi.requests import Request
import bcrypt
from fastapi import UploadFile, File, Form
import shutil
from fastapi.staticfiles import StaticFiles
from fastapi import APIRouter
import openpyxl
from tempfile import NamedTemporaryFile
from io import BytesIO
import zipfile
import unicodedata
import re
import pytz
import locale
# Pillow ekle
from PIL import Image
from fastapi.exception_handlers import RequestValidationError
from fastapi.exception_handlers import request_validation_exception_handler

# Ortam değişkenlerini yükle
load_dotenv()

# FastAPI uygulaması oluştur
app = FastAPI()

# uploads klasörünü statik olarak sun
uploads_dir = os.path.join(os.getcwd(), "uploads")
os.makedirs(uploads_dir, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=uploads_dir), name="uploads")

# CORS ayarları
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Production'da spesifik domainler ekleyin
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# JWT Configy
SECRET_KEY = os.getenv("SECRET_KEY", "113802009Alpha*!")
ALGORITHM = os.getenv("ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", 90))

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# SQLite database connection
def get_db_connection():
    try:
        conn = sqlite3.connect('bandirma.db')
        conn.row_factory = sqlite3.Row  # Sözlük benzeri erişim için
        return conn
    except sqlite3.Error as e:
        print(f"SQLite bağlantı hatası: {e}")
        return None

# Modeller
class Token(BaseModel):
    access_token: str
    token_type: str

class User(BaseModel):
    id: int
    username: str
    role: str

class PasswordChangeRequest(BaseModel):
    old_password: str
    new_password: str

class PasswordChangeOnlyNewRequest(BaseModel):
    new_password: str

class UserCreateRequest(BaseModel):
    username: str
    password: str
    role: str

class KursiyerEkleRequest(BaseModel):
    kurs_durumu: str = ""
    aday_ismi: str = ""
    tel_no: str = ""
    tel_yakin: str = ""
    tc_kimlik: str = ""
    ogrenim_belgesi: Union[int, str] = 0
    adres_belgesi: Union[int, str] = 0
    adli_sicil: Union[int, str] = 0
    ehliyet: Union[int, str] = 0
    kimlik_belgesi: Union[int, str] = 0
    fotograf: Union[int, str] = 0
    basvuru_formu: Union[int, str] = 0
    e_src_kaydi: Union[int, str] = 0
    sinav_ucreti_dahil: Union[int, str] = 0
    alacagi_egitim: str = ""
    devam_egitimi: str = ""
    odeme_durumu: str = ""  
    tutar: Union[float, str] = 0
    tarih_1: str = ""
    odeme_1: Union[float, str] = 0
    tarih_2: str = ""
    odeme_2: Union[float, str] = 0
    tarih_3: str = ""
    odeme_3: Union[float, str] = 0
    tarih_4: str = ""
    odeme_4: Union[float, str] = 0
    tarih_5: str = ""
    odeme_5: Union[float, str] = 0
    tarih_6: str = ""
    odeme_6: Union[float, str] = 0
    kalan: Union[float, str] = 0
    aciklama: str = ""
    evrak_kayit_tarihi: str = ""
    inaktif: Union[int, str] = 0
    arsiv: Union[int, str] = 0
    ogrenim_tarih: str = ""
    adres_tarih: str = ""
    sicil_tarih: str = ""
    ehliyet_tarih: str = ""
    kimlik_tarih: str = ""
    foto_tarih: str = ""
    basvuru_tarih: str = ""
    src_tarih: str = ""

class SinifEkleRequest(BaseModel):
    sinif_isim: str
    kursiyer_sayi: int = 0
    kursiyer_list: list = []

# Kullanıcı doğrulama
def authenticate_user(username: str, password: str):
    conn = get_db_connection()
    if not conn:
        return False
    
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kullanici WHERE kullanici_adi = ?", (username,))
    user = cursor.fetchone()
    conn.close()
    
    if not user:
        return False
    if not pwd_context.verify(password, user["sifre_hash"]):
        return False
    return user

# JWT token oluşturma
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# Token doğrulama
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Oturum süresi doldu, lütfen tekrar giriş yapınız.",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except jwt.ExpiredSignatureError:
        raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
    
    conn = get_db_connection()
    if not conn:
        raise credentials_exception
    
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kullanici WHERE kullanici_adi = ?", (username,))
    user = cursor.fetchone()
    conn.close()
    
    if user is None:
        raise credentials_exception
    return user

# Check bcrypt version compatibility for passlib
if not hasattr(bcrypt, "__about__"):
    import sys
    print(
        "WARNING: Incompatible bcrypt version detected. "
        "Please install 'bcrypt<4.0.0' for passlib compatibility.\n"
        "Run: pip install 'bcrypt<4.0.0'"
    )
    sys.exit(1)

# Helper function to log admin activities
def log_admin_activity(cursor, user: str, aday_ismi: str, degisiklik: str):
    """Log admin activities to the log table"""
    tarih = datetime.now(pytz.timezone("Europe/Istanbul")).strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(
        "INSERT INTO log (user, aday_ismi, degisiklik, tarih) VALUES (?, ?, ?, ?)",
        (user, aday_ismi, degisiklik, tarih)
    )

# Endpoint'ler
@app.post("/token", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    user = authenticate_user(form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Log login activity for all users
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        log_admin_activity(cursor, user["kullanici_adi"], "Sistem", "Giriş yaptı")
        conn.commit()
        conn.close()
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["kullanici_adi"]}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/kullanici/", response_model=List[User])
async def list_kullanici(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, kullanici_adi, rol FROM kullanici")
        rows = cursor.fetchall()
        users = [
            {"id": row["id"], "username": row["kullanici_adi"], "role": row["rol"]}
            for row in rows
        ]
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Kullanıcılar alınamadı: {str(e)}")
    conn.close()
    return users

@app.get("/kullanici/ayarlar")
async def kullanici_ayarlar(current_user: dict = Depends(get_current_user)):
    # Örnek veri, ihtiyaca göre düzenleyin
    return {
        "username": current_user["kullanici_adi"],
        "role": current_user["rol"],
        # ...diğer ayarlar...
    }

@app.post("/kullanici/sifre-degistir")
async def change_password(
    req: PasswordChangeRequest,
    current_user: dict = Depends(get_current_user)
):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    # Eski şifreyi doğrula
    if not pwd_context.verify(req.old_password, current_user["sifre_hash"]):
        conn.close()
        raise HTTPException(status_code=400, detail="Eski şifre yanlış")
    # Yeni şifreyi hashle ve güncelle
    new_hash = pwd_context.hash(req.new_password)
    cursor.execute(
        "UPDATE kullanici SET sifre_hash = ? WHERE kullanici_adi = ?",
        (new_hash, current_user["kullanici_adi"])
    )
    
    # Log şifre değiştirme activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Şifre değiştirdi")
    
    conn.commit()
    conn.close()
    return {"detail": "Şifre başarıyla değiştirildi."}

@app.post("/users/change-password/")
async def change_password_only_new(
    req: PasswordChangeOnlyNewRequest,
    current_user: dict = Depends(get_current_user)
):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    # Yeni şifreyi hashle ve güncelle
    new_hash = pwd_context.hash(req.new_password)
    cursor.execute(
        "UPDATE kullanici SET sifre_hash = ? WHERE kullanici_adi = ?",
        (new_hash, current_user["kullanici_adi"])
    )
    
    # Log şifre değiştirme activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Şifre değiştirdi (Yeni)")
    
    conn.commit()
    conn.close()
    return {"detail": "Parola başarıyla değiştirildi."}

@app.post("/kullanici/ekle")
async def kullanici_ekle(
    req: UserCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kullanici WHERE kullanici_adi = ?", (req.username,))
    if cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Kullanıcı zaten mevcut")
    sifre_hash = pwd_context.hash(req.password)
    cursor.execute(
        "INSERT INTO kullanici (kullanici_adi, sifre_hash, rol) VALUES (?, ?, ?)",
        (req.username, sifre_hash, req.role)
    )
    conn.commit()
    conn.close()
    return {"detail": "Kullanıcı başarıyla eklendi."}

@app.post("/kursiyer/ekle")
async def kursiyer_ekle(
    req: KursiyerEkleRequest,
    current_user: dict = Depends(get_current_user)
):
    # odeme_durumu alanına ödenecek tutarı yaz
    req_dict = req.dict()
    req_dict["odeme_durumu"] = str(req_dict.get("tutar", "")) if req_dict.get("tutar", "") != "" else ""

    # Convert numeric fields to correct types or default to 0
    def to_float(val):
        try:
            return float(val)
        except Exception:
            return 0.0

    def to_int(val):
        try:
            return int(val)
        except Exception:
            return 0

    for key in [
        "ogrenim_belgesi", "adres_belgesi", "adli_sicil", "ehliyet", "kimlik_belgesi",
        "fotograf", "basvuru_formu", "e_src_kaydi"
    ]:
        req_dict[key] = to_int(req_dict.get(key, 0))
    
    # arsiv alanını özel olarak işle
    arsiv_value = req_dict.get("arsiv", 0)
    if arsiv_value in [1, '1', True]:
        req_dict["arsiv"] = 1
    else:
        req_dict["arsiv"] = 0

    for key in [
        "tutar", "odeme_1", "odeme_2", "odeme_3", "odeme_4", "odeme_5", "odeme_6", "kalan"
    ]:
        req_dict[key] = to_float(req_dict.get(key, 0))

    conn = get_db_connection()
    if not conn:
        print("Kursiyer ekleme başarısız: Veritabanı bağlantı hatası")
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        # Eğitim alanlarını büyük harfe çevir
        alacagi_egitim = req_dict.get("alacagi_egitim", "").upper() if req_dict.get("alacagi_egitim", "") else ""
        devam_egitimi = req_dict.get("devam_egitimi", "").upper() if req_dict.get("devam_egitimi", "") else ""
        
        cursor.execute("""
            INSERT INTO kursiyer (
                kurs_durumu, aday_ismi, tel_no, tel_yakin, tc_kimlik,
                ogrenim_belgesi, adres_belgesi, adli_sicil, ehliyet, kimlik_belgesi,
                fotograf, basvuru_formu, e_src_kaydi, sinav_ucreti_dahil, alacagi_egitim, devam_egitimi,
                tutar, tarih_1, odeme_1, tarih_2, odeme_2, tarih_3, odeme_3,
                tarih_4, odeme_4, tarih_5, odeme_5, tarih_6, odeme_6, kalan, aciklama,
                evrak_kayit_tarihi, inaktif, arsiv,
                ogrenim_tarih, adres_tarih, sicil_tarih, ehliyet_tarih, kimlik_tarih, foto_tarih, basvuru_tarih, src_tarih
            ) VALUES (
                ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?, ?, ?
            )
        """, (
            req_dict.get("kurs_durumu", ""),
            req_dict.get("aday_ismi", ""),
            req_dict.get("tel_no", ""),
            req_dict.get("tel_yakin", ""),
            req_dict.get("tc_kimlik", ""),
            req_dict.get("ogrenim_belgesi", 0),
            req_dict.get("adres_belgesi", 0),
            req_dict.get("adli_sicil", 0),
            req_dict.get("ehliyet", 0),
            req_dict.get("kimlik_belgesi", 0),
            req_dict.get("fotograf", 0),
            req_dict.get("basvuru_formu", 0),
            req_dict.get("e_src_kaydi", 0),
            req_dict.get("sinav_ucreti_dahil", 0),
            alacagi_egitim,
            devam_egitimi,
            req_dict.get("tutar", 0),
            req_dict.get("tarih_1", ""),
            req_dict.get("odeme_1", 0),
            req_dict.get("tarih_2", ""),
            req_dict.get("odeme_2", 0),
            req_dict.get("tarih_3", ""),
            req_dict.get("odeme_3", 0),
            req_dict.get("tarih_4", ""),
            req_dict.get("odeme_4", 0),
            req_dict.get("tarih_5", ""),
            req_dict.get("odeme_5", 0),
            req_dict.get("tarih_6", ""),
            req_dict.get("odeme_6", 0),
            req_dict.get("kalan", 0),
            req_dict.get("aciklama", ""),
            req_dict.get("evrak_kayit_tarihi", ""),
            req_dict.get("inaktif", 0),
            req_dict.get("arsiv", 0),
            req_dict.get("ogrenim_tarih", ""),
            req_dict.get("adres_tarih", ""),
            req_dict.get("sicil_tarih", ""),
            req_dict.get("ehliyet_tarih", ""),
            req_dict.get("kimlik_tarih", ""),
            req_dict.get("foto_tarih", ""),
            req_dict.get("basvuru_tarih", ""),
            req_dict.get("src_tarih", "")
        ))
        conn.commit()
        yeni_id = cursor.lastrowid  # --- Eklendi: yeni kursiyer id ---
        
        # Log kursiyer ekleme for admin users
        if current_user["rol"] == "admin":
            aday_ismi = req_dict.get('aday_ismi', '')
            log_admin_activity(cursor, current_user["kullanici_adi"], aday_ismi, "kursiyer eklendi")
            conn.commit()
        
        print(f"Kursiyer başarıyla eklendi: {req_dict.get('aday_ismi', '')} ({req_dict.get('tc_kimlik', '')})")
    except Exception as e:
        print(f"Kursiyer ekleme hatası: {str(e)} - Data: {req_dict}")
        conn.close()
        raise HTTPException(status_code=400, detail=f"Kursiyer eklenemedi: {str(e)}")
    conn.close()
    return {"detail": "Kursiyer başarıyla eklendi.", "id": yeni_id}  # --- Eklendi: id ---

@app.get("/kursiyer/liste")
async def kursiyer_liste(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM kursiyer")
        rows = cursor.fetchall()
        kursiyerler = []
        for row in rows:
            kursiyer = dict(row)
            
            # Tüm string alanları tutarlı hale getir
            for key in kursiyer:
                if isinstance(kursiyer[key], str):
                    kursiyer[key] = kursiyer[key].strip() if kursiyer[key] else ""
                elif kursiyer[key] is None:
                    kursiyer[key] = ""
            
            # inaktif alanını tutarlı hale getir
            if kursiyer.get('inaktif') in [1, '1', True]:
                kursiyer['inaktif'] = 1
            else:
                kursiyer['inaktif'] = 0
                
            # arsiv alanını da normalize et
            if kursiyer.get('arsiv') in [1, '1', True]:
                kursiyer['arsiv'] = 1
            else:
                kursiyer['arsiv'] = 0
                
            kursiyerler.append(kursiyer)
        
        # Debug için log
        print(f"Kursiyer listesi gönderildi: {len(kursiyerler)} kursiyer")
        print(f"Kurs durumları: {list(set(k['kurs_durumu'] for k in kursiyerler))}")
        print(f"Eğitim türleri: {list(set(k['alacagi_egitim'] for k in kursiyerler))}")
        print(f"Inaktif sayısı: {sum(1 for k in kursiyerler if k['inaktif'] == 1)}")
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Kursiyerler alınamadı: {str(e)}")
    conn.close()
    return kursiyerler

@app.put("/kursiyer/guncelle/{kid}")
async def kursiyer_guncelle(kid: str, req: dict, current_user: dict = Depends(get_current_user)):
    print(f"DEBUG: PUT endpoint çağrıldı - kid: {kid}, req: {req}, user: {current_user}")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    field = "id"
    cursor.execute("PRAGMA table_info(kursiyer)")
    columns = [col[1] for col in cursor.fetchall()]
    if "id" not in columns:
        field = "tc_kimlik"
    allowed = set(columns) - {"id"}
    update_fields = []
    values = []

    int_fields = {"ogrenim_belgesi", "adres_belgesi", "adli_sicil", "ehliyet", "kimlik_belgesi",
                  "fotograf", "basvuru_formu", "e_src_kaydi", "sinav_ucreti_dahil"}
    float_fields = {"tutar", "odeme_1", "odeme_2", "odeme_3", "odeme_4", "odeme_5", "odeme_6", "kalan"}

    cursor.execute(f"SELECT * FROM kursiyer WHERE {field} = ?", (kid,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Kursiyer bulunamadı")
    
    print(f"DEBUG: Kursiyer bulundu - row: {dict(row)}")
    print(f"DEBUG: Allowed fields: {allowed}")
    print(f"DEBUG: Request fields: {list(req.keys())}")
    # ...existing code...

    # Log için değişiklikleri topla
    log_degisiklikler = []
    log_odeme_tarih_alanlari = [
        "odeme_1", "odeme_2", "odeme_3", "odeme_4", "odeme_5", "odeme_6",
        "tarih_1", "tarih_2", "tarih_3", "tarih_4", "tarih_5", "tarih_6"
    ]
    is_user = current_user["rol"] == "user"

    # --- EKLENDİ: User ise kalan'ı backend'de hesapla ---
    if is_user:
        # Tutarı ve ödemeleri mevcut row'dan ve req'den al
        try:
            tutar = float(row["tutar"] if row["tutar"] is not None else 0)
        except Exception:
            tutar = 0.0
        toplam_odeme = 0.0
        for i in range(1, 7):
            key = f"odeme_{i}"
            # Eğer yeni değer gönderildiyse onu kullan, yoksa eskiyi kullan
            val = req.get(key, row[key] if key in row.keys() else 0)
            try:
                toplam_odeme += float(val) if val not in ("", None) else 0.0
            except Exception:
                pass
        kalan = round(tutar - toplam_odeme, 2)
        # kalan alanı güncellenecek alanlara eklensin
        req["kalan"] = kalan
        # update_fields'a eklenmesi için allowed'a ekle
        allowed = allowed | {"kalan"}
    # --- SON EKLENDİ ---

    kurs_durumu_yeni = req.get('kurs_durumu', row['kurs_durumu'])
    kurs_durumu_eski = row['kurs_durumu']
    
    # --- EKLENDİ: Özel Durum -> Katiliyor geçişi kontrolü ---
    ozel_durum_gecisi = (kurs_durumu_eski == "Özel Durum" and kurs_durumu_yeni == "Katiliyor")
    
    for k in allowed:
        # Eğer kurs_durumu İptal olacaksa, inaktif ve arsiv alanlarını atla
        if kurs_durumu_yeni == 'İptal' and k in ['inaktif', 'arsiv']:
            continue
        if k in req:
            v = req[k]
            if k in int_fields:
                try:
                    v = int(v)
                except Exception:
                    v = 0
            elif k in float_fields:
                try:
                    v = float(v)
                except Exception:
                    v = 0.0
            elif k == "inaktif":
                # inaktif alanını özel olarak işle
                if v in [1, '1', True]:
                    v = 1
                else:
                    v = 0
            elif k in ["alacagi_egitim", "devam_egitimi"]:
                # Eğitim alanlarını büyük harfe çevir
                v = str(v).upper() if v else ""
            # Eğer kurs_durumu 'İptal' ise ve iptal_tarih alanı varsa, iptal_tarih'i de güncelle
            if k == "kurs_durumu" and v == "İptal" and "iptal_tarih" in req:
                update_fields.append("iptal_tarih = ?")
                values.append(req["iptal_tarih"])
            # Sadece user rolü ödeme/tarih alanlarını değiştirebilir, diğerlerini atla
            # --- DEĞİŞTİ: user rolü sinif alanını değiştiremez ---
            if is_user and (k not in log_odeme_tarih_alanlari and k != "kalan"):
                continue
            if is_user and k == "sinif":
                continue
            # Değişiklik var mı kontrolü
            eski = row[k]
            if str(eski) != str(v):
                update_fields.append(f"{k} = ?")
                values.append(v)
                if is_user and k in log_odeme_tarih_alanlari:
                    log_degisiklikler.append(k)
            elif not is_user:
                update_fields.append(f"{k} = ?")
                values.append(v)
        else:
            # Frontend'den gelmeyen alanlar için eski değeri kullan
            v = row[k]
            update_fields.append(f"{k} = ?")
            values.append(v)

    # --- EKLENDİ: Özel Durum -> Katiliyor geçişi için kurs_durumu değişikliği kontrolü ---
    if not update_fields and "kurs_durumu" in req:
        # Sadece kurs_durumu değişiyorsa ve başka alan yoksa, kurs_durumu'nu ekle
        update_fields.append("kurs_durumu = ?")
        values.append(req["kurs_durumu"])
    
    if not update_fields:
        conn.close()
        raise HTTPException(
            status_code=400,
            detail=f"Güncellenecek veri yok veya alanlar hatalı. Gönderilen alanlar: {list(req.keys())}, izin verilenler: {list(allowed)}"
        )

    # --- Kurs durumu İptal ise inaktif ve arsiv otomatik 1, değilse 0 ---
    if kurs_durumu_yeni == 'İptal':
        if (row['inaktif'] != 1) and ('inaktif = ?' not in update_fields):
            update_fields.append('inaktif = ?')
            values.append(1)
        if (row['arsiv'] != 1) and ('arsiv = ?' not in update_fields):
            update_fields.append('arsiv = ?')
            values.append(1)
    else:
        # Önce varsa eskiyi sil
        for f in ['inaktif = ?', 'arsiv = ?', 'iptal_tarih = ?']:
            if f in update_fields:
                idx = update_fields.index(f)
                update_fields.pop(idx)
                values.pop(idx)
        # Sonra kesinlikle ekle
        update_fields.append('inaktif = ?')
        values.append(0)
        update_fields.append('arsiv = ?')
        values.append(0)
        update_fields.append('iptal_tarih = ?')
        values.append(None)

    values.append(kid)
    try:
        cursor.execute(
            f"UPDATE kursiyer SET {', '.join(update_fields)} WHERE {field} = ?",
            values
        )
        
        # --- EKLENDİ: Özel Durum -> Katiliyor geçişi işlemleri ---
        if ozel_durum_gecisi:
            # 1. Sınav verilerini sil
            cursor.execute("DELETE FROM sinav WHERE kursiyer_id = ?", (kid,))
            
            # 2. Eğitim alanlarını güncelle
            # Devam eğitimi varsa, alacagi_egitim'e taşı ve devam_egitimi'ni temizle
            if row['devam_egitimi'] and str(row['devam_egitimi']).strip():
                cursor.execute(
                    "UPDATE kursiyer SET alacagi_egitim = ?, devam_egitimi = ? WHERE id = ?",
                    (row['devam_egitimi'], "", kid)
                )
            else:
                # Devam eğitimi yoksa, alacagi_egitim'i temizle
                cursor.execute(
                    "UPDATE kursiyer SET alacagi_egitim = ? WHERE id = ?",
                    ("", kid)
                )
        # --- SON EKLENDİ ---
        
        # --- EKLENDİ: Sınıf atandığında otomatik sınav kaydı oluştur ---
        # Eğer sinif alanı güncelleniyorsa ve yeni bir sinif atanıyorsa
        if "sinif" in req and req["sinif"] and req["sinif"].strip():
            # Mevcut sınav kaydı var mı kontrol et
            cursor.execute("SELECT COUNT(*) FROM sinav WHERE kursiyer_id = ?", (kid,))
            sinav_kayit_sayisi = cursor.fetchone()[0]
            
            # Eğer sınav kaydı yoksa yeni kayıt oluştur
            if sinav_kayit_sayisi == 0:
                # Kursiyer bilgilerini al
                cursor.execute("SELECT alacagi_egitim, devam_egitimi FROM kursiyer WHERE id = ?", (kid,))
                kursiyer_bilgi = cursor.fetchone()
                
                # Sınav kaydı oluştur
                cursor.execute("""
                    INSERT INTO sinav (kursiyer_id, alacagi_egitim, devam_egitimi) 
                    VALUES (?, ?, ?)
                """, (kid, kursiyer_bilgi['alacagi_egitim'] if kursiyer_bilgi else "", 
                      kursiyer_bilgi['devam_egitimi'] if kursiyer_bilgi else ""))
        # --- SON EKLENDİ ---
        # Log kaydı ekle (sadece user rolü ve ödeme/tarih alanı değiştiyse)
        if is_user and log_degisiklikler:
            aday_ismi = row["aday_ismi"]
            degisiklik = ", ".join(log_degisiklikler) + " değiştirildi"
            # --- DEĞİŞTİ: Türkiye saatine göre tarih ---
            tarih = datetime.now(pytz.timezone("Europe/Istanbul")).strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute(
                "INSERT INTO log (user, aday_ismi, degisiklik, tarih) VALUES (?, ?, ?, ?)",
                (current_user["kullanici_adi"], aday_ismi, degisiklik, tarih)
            )
        
        # Log kursiyer düzenleme for all users
        aday_ismi = row["aday_ismi"]
        log_admin_activity(cursor, current_user["kullanici_adi"], aday_ismi, "kursiyer düzenlendi")
        
        # Check if kursiyer was made inactive (only for admin)
        if current_user["rol"] == "admin" and "inaktif" in req and req["inaktif"] == 1 and not (row["inaktif"] == 1 or row["inaktif"] == '1' or row["inaktif"] == True):
            log_admin_activity(cursor, current_user["kullanici_adi"], aday_ismi, "kursiyer inaktif edildi")
        
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Güncelleme hatası: {str(e)}")
    conn.close()
    return {"detail": "Kursiyer başarıyla güncellendi."}

@app.delete("/kullanici/{user_id}/")
async def delete_kullanici(user_id: int, current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    if current_user["id"] == user_id:
        raise HTTPException(status_code=400, detail="Kendi hesabınızı silemezsiniz")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kullanici WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    cursor.execute("DELETE FROM kullanici WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"detail": "Kullanıcı başarıyla silindi."}

@app.get("/kursiyer/ara")
async def kursiyer_ara(
    type: str = Query(..., pattern="^(isim|tc)$"),
    value: str = Query(...),
    current_user: dict = Depends(get_current_user)
):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        if type == "isim":
            cursor.execute("SELECT * FROM kursiyer WHERE LOWER(aday_ismi) LIKE LOWER(?)", (f"%{value}%",))
        elif type == "tc":
            cursor.execute("SELECT * FROM kursiyer WHERE tc_kimlik LIKE ?", (f"%{value}%",))
        else:
            conn.close()
            raise HTTPException(status_code=400, detail="Geçersiz arama türü")
        rows = cursor.fetchall()
        kursiyerler = []
        for row in rows:
            kursiyer = dict(row)
            
            # Tüm string alanları tutarlı hale getir
            for key in kursiyer:
                if isinstance(kursiyer[key], str):
                    kursiyer[key] = kursiyer[key].strip() if kursiyer[key] else ""
                elif kursiyer[key] is None:
                    kursiyer[key] = ""
            
            # inaktif alanını tutarlı hale getir
            if kursiyer.get('inaktif') in [1, '1', True]:
                kursiyer['inaktif'] = 1
            else:
                kursiyer['inaktif'] = 0
                
            # arsiv alanını da normalize et
            if kursiyer.get('arsiv') in [1, '1', True]:
                kursiyer['arsiv'] = 1
            else:
                kursiyer['arsiv'] = 0
                
            kursiyerler.append(kursiyer)
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Kursiyer arama hatası: {str(e)}")
    conn.close()
    return kursiyerler

@app.post("/api/sinif")
async def sinif_ekle(
    req: SinifEkleRequest,
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO sinif (sinif_isim, kursiyer_sayi, kursiyer_list) VALUES (?, ?, ?)",
            (
                req.sinif_isim,
                req.kursiyer_sayi,
                str(req.kursiyer_list),  
            )
        )
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Sınıf eklenemedi: {str(e)}")
    conn.close()
    return {"detail": "Sınıf başarıyla eklendi."}

@app.get("/api/siniflar")
async def sinif_listele(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM sinif")
        rows = cursor.fetchall()
        siniflar = [dict(row) for row in rows]
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Sınıflar alınamadı: {str(e)}")
    conn.close()
    return siniflar

@app.delete("/api/sinif/{sinif_id}")
async def sinif_sil(sinif_id: int, current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM sinif WHERE id = ?", (sinif_id,))
    sinif = cursor.fetchone()
    if not sinif:
        conn.close()
        raise HTTPException(status_code=404, detail="Sınıf bulunamadı")
    cursor.execute("DELETE FROM sinif WHERE id = ?", (sinif_id,))
    conn.commit()
    conn.close()
    return {"detail": "Sınıf başarıyla silindi."}

@app.get("/api/belge/{kursiyer_id}")
async def belge_getir(kursiyer_id: str, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return {}
    return dict(row)

@app.post("/api/belge/yukle/{kursiyer_id}")
async def belge_yukle(
    kursiyer_id: str,
    belge_tipi: str = Form(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    allowed_keys = [
        "foto", "ogrenim_belgesi", "adres_belgesi", "adli_sicil",
        "ehliyet", "kimlik_belgesi", "basvuru_formu"
    ]
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kursiyer WHERE id = ? OR tc_kimlik = ?", (kursiyer_id, kursiyer_id))
    kursiyer = cursor.fetchone()
    if not kursiyer:
        conn.close()
        raise HTTPException(status_code=404, detail="Kursiyer bulunamadı")
    aday_ismi = (kursiyer["aday_ismi"] or "").strip()
    alacagi_egitim = (kursiyer["alacagi_egitim"] or "").strip().lower()
    uploads_dir = os.path.join(os.getcwd(), "uploads")
    klasor_adi = f"{aday_ismi} - {alacagi_egitim}".strip(" -")
    klasor_path = os.path.join(uploads_dir, klasor_adi)
    os.makedirs(klasor_path, exist_ok=True)
    # Dosya uzantısı ve adı büyük harfe çevrilsin
    ext = os.path.splitext(file.filename)[1].upper()
    # Sadece jpeg/jpg/png/pdf uzantılarını kabul et
    allowed_exts = {".JPEG", ".JPG", ".PNG", ".PDF"}
    if ext not in allowed_exts:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Yalnızca jpeg, jpg, png veya pdf dosyaları yüklenebilir. Dosya uzantısı: {ext}")
    # Dosya adını normalize et
    def normalize_filename(name):
        name = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
        name = re.sub(r'[^A-Za-z0-9_\-\. ]+', '', name)
        return name.strip().replace(' ', '_').upper()
    if belge_tipi == "foto":
        dosya_adi = f"FOTOGRAF{ext}"
    else:
        dosya_adi = f"{normalize_filename(belge_tipi)}{ext}"
    dosya_path = os.path.join(klasor_path, dosya_adi)
    with open(dosya_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # --- DPI Correction for images ---
    if ext in [".JPEG", ".JPG", ".PNG"]:
        try:
            img = Image.open(dosya_path)
            img.save(dosya_path, dpi=(400, 400))
        except Exception as e:
            print(f"DPI ayarlanamadı: {e}")
    # Dosya yolunu uploads/... ile başlat
    rel_path = os.path.relpath(dosya_path, os.getcwd())
    if not rel_path.startswith("uploads"):
        rel_path = os.path.join("uploads", os.path.relpath(dosya_path, uploads_dir))
    kursiyer_real_id = kursiyer["id"]
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer_real_id,))
    mevcut = cursor.fetchone()
    if mevcut:
        cursor.execute(
            f"UPDATE belge SET {belge_tipi} = ? WHERE kursiyer_id = ?",
            (rel_path, kursiyer_real_id)
        )
    else:
        fields = ["kursiyer_id"] + allowed_keys
        values = [kursiyer_real_id] + [rel_path if k == belge_tipi else None for k in allowed_keys]
        placeholders = ",".join(["?"] * len(fields))
        cursor.execute(
            f"INSERT INTO belge ({','.join(fields)}) VALUES ({placeholders})",
            values
        )
    conn.commit()
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer_real_id,))
    belge_row = cursor.fetchone()
    conn.close()
    return dict(belge_row) if belge_row else {}

@app.post("/yedek/kursiyer-yukle")
async def yedek_kursiyer_yukle(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    # Save uploaded file to a temp file
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    errors = []
    imported = 0
    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Get kursiyer table columns
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(kursiyer)")
        kursiyer_columns = [col[1] for col in cursor.fetchall()]
        # Only use columns that exist in kursiyer table
        valid_headers = [h for h in headers if h in kursiyer_columns]
        if not valid_headers:
            raise Exception("Excel sütun başlıkları kursiyer tablosu ile eşleşmiyor.")
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {h: (cell.value if cell.value is not None else "") for h, cell in zip(headers, row)}
            # Sadece aday_ismi doluysa ekle
            aday_ismi = row_data.get("aday_ismi", "")
            if not aday_ismi or str(aday_ismi).strip() == "":
                continue
            # Only keep valid columns except 'kalan'
            insert_data = {k: row_data[k] for k in valid_headers if k != "kalan"}
            
            # inaktif alanını özel olarak işle
            if "inaktif" in insert_data:
                inaktif_value = insert_data["inaktif"]
                if inaktif_value in [1, '1', True]:
                    insert_data["inaktif"] = 1
                else:
                    insert_data["inaktif"] = 0
            # Telefon ve TC kimlik için değerleri daima string olarak al
            for key in ["tel_no", "tel_yakin", "tc_kimlik"]:
                if key in insert_data:
                    val = insert_data[key]
                    if val is None:
                        val = ""
                    # Excel'den gelen değer sayıysa stringe çevir
                    if isinstance(val, (int, float)):
                        # Ondalık kısmı yoksa tam sayı gibi göster
                        if isinstance(val, float) and val.is_integer():
                            val = str(int(val))
                        else:
                            val = str(val)
                    else:
                        val = str(val)
                    insert_data[key] = val
            
            # Eğitim alanlarını büyük harfe çevir
            for key in ["alacagi_egitim", "devam_egitimi"]:
                if key in insert_data:
                    val = insert_data[key]
                    if val is not None and str(val).strip():
                        insert_data[key] = str(val).upper()
                    else:
                        insert_data[key] = ""
            # Kalanı hesapla
            try:
                tutar = float(row_data.get("tutar", 0) or 0)
            except Exception:
                tutar = 0.0
            toplam_odeme = 0.0
            for odeme_key in ["odeme_1", "odeme_2", "odeme_3", "odeme_4", "odeme_5", "odeme_6"]:
                try:
                    toplam_odeme += float(row_data.get(odeme_key, 0) or 0)
                except Exception:
                    pass
            kalan = round(tutar - toplam_odeme, 2)
            # Eğer kursiyer tablosunda 'kalan' sütunu varsa, insert_data'ya ekle
            if "kalan" in kursiyer_columns:
                insert_data["kalan"] = kalan
            # Try to insert
            try:
                insert_cols = list(insert_data.keys())
                placeholders = ",".join(["?"] * len(insert_cols))
                columns_str = ",".join(insert_cols)
                values = [insert_data[k] for k in insert_cols]
                cursor.execute(
                    f"INSERT INTO kursiyer ({columns_str}) VALUES ({placeholders})",
                    values
                )
                imported += 1
            except Exception as e:
                errors.append(f"{idx}. satır: {str(e)}")
        conn.commit()
        conn.close()
    except Exception as e:
        errors.append(f"Dosya okunamadı veya işlenemedi: {str(e)}")
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
    return {"imported": imported, "errors": errors}

@app.get("/yedek/db-indir")
async def yedek_db_indir(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    db_path = os.path.join(os.getcwd(), "bandirma.db")
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Veritabanı dosyası bulunamadı")
    return FileResponse(
        db_path,
        media_type="application/octet-stream",
        filename="bandirma.db"
    )

@app.get("/api/sinav/{kursiyer_id}")
async def get_sinav(kursiyer_id: str, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM sinav WHERE kursiyer_id = ?", (kursiyer_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return {}
    return dict(row)

@app.post("/api/sinav/{kursiyer_id}")
async def upsert_sinav(
    kursiyer_id: str,
    req: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    # Sınav tablosu kolonları
    cursor.execute("PRAGMA table_info(sinav)")
    sinav_columns = [col[1] for col in cursor.fetchall()]
    allowed = set(sinav_columns) - {"id"}
    cursor.execute("SELECT * FROM sinav WHERE kursiyer_id = ?", (kursiyer_id,))
    mevcut = cursor.fetchone()
    data = {k: req[k] for k in allowed if k in req}
    data["kursiyer_id"] = int(kursiyer_id)
    
    yazili_keys = [f"sinav_{i+1}" for i in range(4)]
    uygulama_keys = [f"uygulama_{i+1}" for i in range(4)]
    devam_yazili_keys = [f"devam_sinav_{i+1}" for i in range(4)]
    devam_uygulama_keys = [f"devam_uygulama_{i+1}" for i in range(4)]
    yazili_tarih_keys = [f"sinav_{i+1}_tarih" for i in range(4)]
    uygulama_tarih_keys = [f"uygulama_{i+1}_tarih" for i in range(4)]
    devam_yazili_tarih_keys = [f"devam_sinav_{i+1}_tarih" for i in range(4)]
    devam_uygulama_tarih_keys = [f"devam_uygulama_{i+1}_tarih" for i in range(4)]

    # --- Puanları clamp et ---
    for k in yazili_keys + devam_yazili_keys:
        if k in data and data[k] is not None and data[k] != "":
            try:
                puan = float(data[k])
                if puan < 0:
                    puan = 0
                if puan > 100:
                    puan = 100
                data[k] = int(puan)
            except Exception:
                data[k] = None

    # --- Tarih alanlarını normalize et ("" ise None yap) ---
    for key in yazili_tarih_keys + uygulama_tarih_keys + devam_yazili_tarih_keys + devam_uygulama_tarih_keys:
        if key in data and (data[key] is None or data[key] == ""):
            data[key] = None

    # --- Devam eğitimi için: Eğer not girilmemişse tarihi temizle ---
    for idx, k in enumerate(devam_yazili_keys):
        if str(data.get(k, "")) in ("", "None", "null") and data.get(devam_yazili_tarih_keys[idx], None) not in (None, "", "null"):
            data[devam_yazili_tarih_keys[idx]] = None
    
    for idx, k in enumerate(devam_uygulama_keys):
        if str(data.get(k, "")) in ("", "None", "null") and data.get(devam_uygulama_tarih_keys[idx], None) not in (None, "", "null"):
            data[devam_uygulama_tarih_keys[idx]] = None

    try:
        # Sınav ücreti dahil kontrolü ve güncelleme
        kursiyer_bilgisi = cursor.execute("SELECT sinav_ucreti_dahil, kurs_durumu, devam_egitimi FROM kursiyer WHERE id = ?", (kursiyer_id,)).fetchone()
        
        if kursiyer_bilgisi:
            sinav_ucreti_dahil = kursiyer_bilgisi["sinav_ucreti_dahil"]
            kurs_durumu = kursiyer_bilgisi["kurs_durumu"]
            devam_egitimi = kursiyer_bilgisi["devam_egitimi"]
            
            # İlk yazılı sınav puanı girildiğinde kontrol
            yazili_girilmis = any(str(data.get(k, "")) not in ("", "None", "null") for k in yazili_keys)
            devam_yazili_girilmis = any(str(data.get(k, "")) not in ("", "None", "null") for k in devam_yazili_keys)
            
            # Eğer ilk yazılı sınav puanı girildiyse
            if (yazili_girilmis or devam_yazili_girilmis) and sinav_ucreti_dahil == 1:
                # Kurs durumu "Fark" ise ve devam eğitimi varsa, sinav_ucreti_dahil'i tekrar 1 yap
                if kurs_durumu == "Fark" and devam_egitimi and str(devam_egitimi).strip() != "":
                    # Devam eğitiminde ilk yazılı puanı girildiğinde sinav_ucreti_dahil = 0 yap
                    if devam_yazili_girilmis:
                        cursor.execute("UPDATE kursiyer SET sinav_ucreti_dahil = 0 WHERE id = ?", (kursiyer_id,))
                else:
                    # Normal durumda ilk yazılı puanı girildiğinde sinav_ucreti_dahil = 0 yap
                    cursor.execute("UPDATE kursiyer SET sinav_ucreti_dahil = 0 WHERE id = ?", (kursiyer_id,))
        
        if mevcut:
            update_fields = [f"{k} = ?" for k in data if k != "kursiyer_id"]
            values = [data[k] for k in data if k != "kursiyer_id"]
            values.append(kursiyer_id)
            cursor.execute(
                f"UPDATE sinav SET {', '.join(update_fields)} WHERE kursiyer_id = ?",
                values
            )
        else:
            insert_cols = list(data.keys())
            placeholders = ",".join(["?"] * len(insert_cols))
            columns_str = ",".join(insert_cols)
            values = [data[k] for k in insert_cols]
            cursor.execute(
                f"INSERT INTO sinav ({columns_str}) VALUES ({placeholders})",
                values
            )

        # --- EKLENDİ: Uygulama sınavları kontrolü ile kurs_durumu güncelle ---
        uygulama_vals = [data.get(f"uygulama_{i+1}", "") for i in range(4)]
        uygulama_hepsi_bos = all((v is None or v == "") for v in uygulama_vals)
        uygulama_bir_kaldi = any(v == "Kaldı" for v in uygulama_vals)
        cursor.execute("SELECT kurs_durumu FROM kursiyer WHERE id = ?", (kursiyer_id,))
        kursiyer_row = cursor.fetchone()
        if kursiyer_row:
            mevcut_kurs_durumu = kursiyer_row["kurs_durumu"]
            if (mevcut_kurs_durumu == "Fark" or mevcut_kurs_durumu != "Katiliyor") and (uygulama_hepsi_bos or uygulama_bir_kaldi):
                cursor.execute("UPDATE kursiyer SET kurs_durumu = 'Katiliyor' WHERE id = ?", (kursiyer_id,))

        # --- EKLENDİ: Kursiyerin devam_egitimi varsa ve uygulama_X "Geçti" ise kurs_durumu = "Fark" yap ---
        cursor.execute("SELECT devam_egitimi FROM kursiyer WHERE id = ?", (kursiyer_id,))
        kursiyer_row = cursor.fetchone()
        devam_egitimi_var = kursiyer_row and kursiyer_row["devam_egitimi"] and str(kursiyer_row["devam_egitimi"]).strip() != ""
        uygulama_gecti = any(
            (data.get(f"uygulama_{i+1}", "") == "Geçti") for i in range(4)
        )
        if devam_egitimi_var and uygulama_gecti:
            cursor.execute("UPDATE kursiyer SET kurs_durumu = 'Fark' WHERE id = ?", (kursiyer_id,))
            # Fark durumuna geçiş yapıldığında sinav_ucreti_dahil'i 1 yap
            cursor.execute("UPDATE kursiyer SET sinav_ucreti_dahil = 1 WHERE id = ?", (kursiyer_id,))
        # --- SON EKLENDİ ---

        # --- YENİ: SHDA kontrolü ---
        # 1) Yazılı 1-4 tamamen 70'ten küçükse (ve sayı ise)
        yazili_vals = [data.get(f"sinav_{i+1}", None) for i in range(4)]
        yazili_hepsi_kucuk_70 = all(
            (v is not None and str(v) != "" and isinstance(v, (int, float)) and float(v) < 70)
            for v in yazili_vals
        )
        # 2) Uygulama 1-4 tamamen 'Kaldı' veya 'Girmedi' ise (başarısız)
        uygulama_vals = [data.get(f"uygulama_{i+1}", "") for i in range(4)]
        uygulama_hepsi_basarisiz = all(v in ["Kaldı", "Girmedi"] for v in uygulama_vals)
        # 3) Yazılılardan herhangi biri 70 ve üzeri (geçmiş)
        yazili_herhangi_gecmis = any(
            (v is not None and str(v) != "" and isinstance(v, (int, float)) and float(v) >= 70)
            for v in yazili_vals
        )
        
        # Mevcut kursiyer durumu kontrolü
        cursor.execute("SELECT kurs_durumu, inaktif, arsiv FROM kursiyer WHERE id = ?", (kursiyer_id,))
        kursiyer_row = cursor.fetchone()
        mevcut_kurs_durumu = kursiyer_row["kurs_durumu"] if kursiyer_row else None
        mevcut_inaktif = kursiyer_row["inaktif"] if kursiyer_row else 0
        mevcut_arsiv = kursiyer_row["arsiv"] if kursiyer_row else 0
        
        # SHDA koşulları kontrolü - DOĞRU MANTIK
        # 1) Yazılı 1-4 hepsi 70 altında VEYA
        # 2) Yazılılardan herhangi biri geçmiş VE uygulamadan hepsi başarısız
        if yazili_hepsi_kucuk_70 or (yazili_herhangi_gecmis and uygulama_hepsi_basarisiz):
            # SHDA durumunda mutlaka inaktif = 1 ve arsiv = 1 yap
            cursor.execute("UPDATE kursiyer SET kurs_durumu = 'SHDA', inaktif = 1, arsiv = 1 WHERE id = ?", (kursiyer_id,))
        else:
            # SHDA koşulu artık sağlanmıyorsa, kurs_durumu ne olursa olsun, inaktif veya arsiv 1 ise 0'a çek
            if mevcut_kurs_durumu == "SHDA":
                cursor.execute("UPDATE kursiyer SET kurs_durumu = 'Katiliyor' WHERE id = ?", (kursiyer_id,))
            if mevcut_inaktif in [1, '1', True]:
                cursor.execute("UPDATE kursiyer SET inaktif = 0 WHERE id = ?", (kursiyer_id,))
            if mevcut_arsiv in [1, '1', True]:
                cursor.execute("UPDATE kursiyer SET arsiv = 0 WHERE id = ?", (kursiyer_id,))
            # Ekstra güvenlik: inaktif veya arsiv NULL ise 0'a çek
            if mevcut_inaktif in (None, "", "null"): 
                cursor.execute("UPDATE kursiyer SET inaktif = 0 WHERE id = ?", (kursiyer_id,))
            if mevcut_arsiv in (None, "", "null"): 
                cursor.execute("UPDATE kursiyer SET arsiv = 0 WHERE id = ?", (kursiyer_id,))

        # Log exam result entry for admin users
        if current_user["rol"] == "admin":
            cursor.execute("SELECT aday_ismi FROM kursiyer WHERE id = ?", (kursiyer_id,))
            kursiyer_name_row = cursor.fetchone()
            if kursiyer_name_row:
                aday_ismi = kursiyer_name_row["aday_ismi"]
                log_admin_activity(cursor, current_user["kullanici_adi"], aday_ismi, "Yazılı/Uygulama Sınav Sonucu Girildi")

        # --- KURSU TAMAMLAMA KONTROLÜ ---
        # 1) Kursiyerin devam_egitimi boş mu?
        cursor.execute("SELECT devam_egitimi FROM kursiyer WHERE id = ?", (kursiyer_id,))
        kursiyer_row = cursor.fetchone()
        devam_egitimi_bos = not (kursiyer_row and kursiyer_row["devam_egitimi"] and str(kursiyer_row["devam_egitimi"]).strip())
        # 2) Uygulama sınavı 1-4'ten herhangi biri 'Geçti' mi?
        uygulama_gecti_var = False
        for i in range(4):
            val = data.get(f"uygulama_{i+1}", None)
            if val == "Geçti":
                uygulama_gecti_var = True
        # 3) Mevcut kursiyer durumu ve alanları
        cursor.execute("SELECT kurs_durumu, inaktif, arsiv FROM kursiyer WHERE id = ?", (kursiyer_id,))
        kursiyer_row = cursor.fetchone()
        mevcut_kurs_durumu = kursiyer_row["kurs_durumu"] if kursiyer_row else None
        mevcut_inaktif = kursiyer_row["inaktif"] if kursiyer_row else 0
        mevcut_arsiv = kursiyer_row["arsiv"] if kursiyer_row else 0
        if devam_egitimi_bos and uygulama_gecti_var:
            # Kursu tamamladı
            cursor.execute("UPDATE kursiyer SET kurs_durumu = ?, inaktif = 1, arsiv = 1 WHERE id = ?", ("Kursu Tamamladi!", kursiyer_id))
        elif mevcut_kurs_durumu == "Kursu Tamamladi!" and not uygulama_gecti_var:
            # Artık tamamlamış değil
            cursor.execute("UPDATE kursiyer SET kurs_durumu = ?, inaktif = 0, arsiv = 0 WHERE id = ?", ("Katiliyor", kursiyer_id))
        # --- Uygulama Geçti -> başka bir şeye dönerse tarihini sil ---
        for i in range(4):
            uygulama_key = f"uygulama_{i+1}"
            tarih_key = f"uygulama_{i+1}_tarih"
            # Eğer mevcutta Geçti idi ve şimdi başka bir şey olduysa tarihi sil
            if mevcut and mevcut[uygulama_key] == "Geçti" and data.get(uygulama_key, "") != "Geçti":
                data[tarih_key] = None
        # --- SONU ---

        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Sınav kaydı hatası: {str(e)}")
    conn.close()
    return {"detail": "Sınav bilgisi kaydedildi."}

@app.get("/api/sinavlar/filtreli")
async def sinavlar_filtreli(
    filtre: str = Query(..., pattern="^(yazilidan_kalanlar|yazilidan_gecenler|kursu_tamamlayanlar|sinav_listesi)$"),
    current_user: dict = Depends(get_current_user)
):
    """
    Filtreye göre kursiyer ve sınav bilgilerini döndürür.
    Dönüş: [{aday_ismi, tel_no, tc_kimlik, alacagi_egitim, sinif, ...sinav alanları}]
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            k.id as kursiyer_id, k.aday_ismi, k.tel_no, k.tc_kimlik, k.alacagi_egitim, k.devam_egitimi, k.kurs_durumu, k.sinif, k.sinav_ucreti_dahil,
            s.*
        FROM kursiyer k
        LEFT JOIN sinav s ON k.id = s.kursiyer_id
    """)
    rows = cursor.fetchall()
    result = []
    for row in rows:
        row = dict(row)
        # SINIFI OLMAYANLARI ATLAMAK İÇİN:
        if not row.get("sinif") or str(row.get("sinif")).strip() == "":
            continue
        
        # SHDA durumundaki kursiyerleri atlamak için:
        if row.get("kurs_durumu") == "SHDA":
            continue
        
        # Kursiyer durumuna göre hangi eğitimin sınav notlarını kontrol edeceğimizi belirle
        kurs_durumu = row.get("kurs_durumu", "")
        alacagi_egitim = row.get("alacagi_egitim", "")
        devam_egitimi = row.get("devam_egitimi", "")
        
        # Devam eğitimi notları varsa, hangi durumda olursa olsun bunları kontrol et
        if devam_egitimi and str(devam_egitimi).strip() != "":
            # Devam eğitimi sınav notlarını kullan
            yazili = [row.get(f"devam_sinav_{i+1}") for i in range(4)]
            uygulama = [row.get(f"devam_uygulama_{i+1}") for i in range(4)]
        else:
            # Normal alacagi_egitim sınav notlarını kullan
            yazili = [row.get(f"sinav_{i+1}") for i in range(4)]
            uygulama = [row.get(f"uygulama_{i+1}") for i in range(4)]
        
        # Yazılı sınavı kontrolü
        yazili_girilmis = any(puan is not None and str(puan) != "" for puan in yazili)
        yazili_gecen = any(
            (puan is not None and str(puan) != "" and float(puan) >= 70)
            for puan in yazili
        )
        # Yazılıda en az bir sınavda kalmış ama henüz 4/4 başarısız değil
        yazili_devam_eden = any(
            (puan is not None and str(puan) != "" and float(puan) < 70)
            for puan in yazili
        ) and not yazili_gecen and yazili_girilmis
        
        # Uygulama sınavı kontrolü
        uygulama_girilmis = any(puan is not None and str(puan) != "" for puan in uygulama)
        uygulama_gecen = any(puan == "Geçti" for puan in uygulama)
        # Uygulamada en az bir sınavda kalmış ama henüz 4/4 başarısız değil
        uygulama_devam_eden = any(
            (puan is not None and str(puan) != "" and puan != "Geçti")
            for puan in uygulama
        ) and not uygulama_gecen and uygulama_girilmis
        
        # Henüz hiç sınav notu girilmemiş mi kontrol et
        hic_sinav_girilmemis = not yazili_girilmis and not uygulama_girilmis
        
        # Kursu tamamlayan = Uygulamadan geçen
        kursu_tamamlayan = uygulama_gecen

        # Filtre mantığı:
        if filtre == "yazilidan_kalanlar" and yazili_devam_eden:
            result.append(row)
        elif filtre == "yazilidan_gecenler" and yazili_gecen and not kursu_tamamlayan:
            result.append(row)
        elif filtre == "kursu_tamamlayanlar" and kursu_tamamlayan:
            result.append(row)
        elif filtre == "sinav_listesi" and (yazili_devam_eden or uygulama_devam_eden or hic_sinav_girilmemis):
            result.append(row)
    conn.close()
    filtered = []
    for row in result:
        # Kursiyer durumuna göre hangi eğitimin bilgilerini döndüreceğimizi belirle
        kurs_durumu = row.get("kurs_durumu", "")
        alacagi_egitim = row.get("alacagi_egitim", "")
        devam_egitimi = row.get("devam_egitimi", "")
        
        if devam_egitimi and str(devam_egitimi).strip() != "":
            # Devam eğitimi bilgilerini döndür
            filtered.append({
                "aday_ismi": row.get("aday_ismi"),
                "tel_no": row.get("tel_no"),
                "tc_kimlik": row.get("tc_kimlik"),
                "alacagi_egitim": devam_egitimi,  # Devam eğitimini göster
                "sinif": row.get("sinif"),
                "sinav_ucreti_dahil": row.get("sinav_ucreti_dahil"),
                "sinav_1": row.get("devam_sinav_1"),
                "sinav_2": row.get("devam_sinav_2"),
                "sinav_3": row.get("devam_sinav_3"),
                "sinav_4": row.get("devam_sinav_4"),
                "uygulama_1": row.get("devam_uygulama_1"),
                "uygulama_2": row.get("devam_uygulama_2"),
                "uygulama_3": row.get("devam_uygulama_3"),
                "uygulama_4": row.get("devam_uygulama_4"),
            })
        else:
            # Normal alacagi_egitim bilgilerini döndür
            filtered.append({
                "aday_ismi": row.get("aday_ismi"),
                "tel_no": row.get("tel_no"),
                "tc_kimlik": row.get("tc_kimlik"),
                "alacagi_egitim": alacagi_egitim,
                "sinif": row.get("sinif"),
                "sinav_ucreti_dahil": row.get("sinav_ucreti_dahil"),
                "sinav_1": row.get("sinav_1"),
                "sinav_2": row.get("sinav_2"),
                "sinav_3": row.get("sinav_3"),
                "sinav_4": row.get("sinav_4"),
                "uygulama_1": row.get("uygulama_1"),
                "uygulama_2": row.get("uygulama_2"),
                "uygulama_3": row.get("uygulama_3"),
                "uygulama_4": row.get("uygulama_4"),
            })
    
    # Türkçe karakterleri dönüştürme fonksiyonu
    def turkish_sort_key(name):
        # Türkçe karakterleri ASCII karşılıklarına çevir
        turkish_chars = {
            'Ç': 'C', 'ç': 'c',
            'Ğ': 'G', 'ğ': 'g', 
            'İ': 'I', 'ı': 'i',
            'Ö': 'O', 'ö': 'o',
            'Ş': 'S', 'ş': 's',
            'Ü': 'U', 'ü': 'u'
        }
        
        converted = ""
        for char in name:
            converted += turkish_chars.get(char, char)
        
        return converted.lower()
    
    # Alfabetik sıralama ekle
    filtered.sort(key=lambda x: turkish_sort_key(x["aday_ismi"]))
    return filtered

@app.get("/api/loglar")
async def loglari_getir(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT l.user, l.aday_ismi, l.degisiklik, l.tarih
            FROM log l
            ORDER BY l.tarih DESC
        """)
        rows = cursor.fetchall()
        loglar = [dict(row) for row in rows]
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Loglar alınamadı: {str(e)}")
    conn.close()
    return loglar

@app.get("/api/rapor/tamamlayanlar")
async def rapor_tamamlayanlar(
    baslangic_tarihi: str = Query(...),
    bitis_tarihi: str = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Belirtilen tarih aralığında kursu tamamlayan kursiyerleri listeler.
    Kursu tamamlayan: uygulama_X veya devam_uygulama_X alanlarından herhangi biri "Geçti" olan
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    
    # Log rapor alma activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Tamamlayanlar raporu aldı")
    
    try:
        cursor.execute("""
            SELECT DISTINCT
                k.id,
                k.aday_ismi,
                k.tc_kimlik,
                k.tel_no,
                k.sinif,
                k.alacagi_egitim,
                k.devam_egitimi
            FROM kursiyer k
            INNER JOIN sinav s ON k.id = s.kursiyer_id
            WHERE (
                s.uygulama_1 = 'Geçti' OR s.uygulama_2 = 'Geçti' OR 
                s.uygulama_3 = 'Geçti' OR s.uygulama_4 = 'Geçti' OR
                s.devam_uygulama_1 = 'Geçti' OR s.devam_uygulama_2 = 'Geçti' OR
                s.devam_uygulama_3 = 'Geçti' OR s.devam_uygulama_4 = 'Geçti'
            )
            AND (
                (s.uygulama_1_tarih BETWEEN ? AND ?) OR
                (s.uygulama_2_tarih BETWEEN ? AND ?) OR
                (s.uygulama_3_tarih BETWEEN ? AND ?) OR
                (s.uygulama_4_tarih BETWEEN ? AND ?) OR
                (s.devam_uygulama_1_tarih BETWEEN ? AND ?) OR
                (s.devam_uygulama_2_tarih BETWEEN ? AND ?) OR
                (s.devam_uygulama_3_tarih BETWEEN ? AND ?) OR
                (s.devam_uygulama_4_tarih BETWEEN ? AND ?)
            )
            ORDER BY k.aday_ismi
        """, (
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi,
            baslangic_tarihi, bitis_tarihi
        ))
        rows = cursor.fetchall()
        result = []
        for row in rows:
            result.append({
                "id": row["id"],
                "aday_ismi": row["aday_ismi"],
                "tc_kimlik": row["tc_kimlik"],
                "tel_no": row["tel_no"],
                "sinif": row["sinif"],
                "alacagi_egitim": row["alacagi_egitim"],
                "devam_egitimi": row["devam_egitimi"]
            })
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Rapor alınamadı: {str(e)}")
    conn.close()
    return result

@app.get("/api/rapor/eksik-evrak")
async def rapor_eksik_evrak(current_user: dict = Depends(get_current_user)):
    """
    kurs_durumu = "Beklemede" olan ve eksik evrakı olan VEYA ödemesi tam olmayan kursiyerleri listeler.
    Eksik evrak: ogrenim_belgesi, adres_belgesi, adli_sicil, ehliyet, 
    kimlik_belgesi, fotograf, basvuru_formu, e_src_kaydi alanlarından herhangi biri 0 olan
    Eksik ödeme: kalan > 0 olan
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    
    # Log rapor alma activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Eksik evrak raporu aldı")
    
    try:
        cursor.execute("""
            SELECT 
                k.id,
                k.aday_ismi,
                k.tc_kimlik,
                k.tel_no,
                k.sinif,
                k.alacagi_egitim,
                k.ogrenim_belgesi,
                k.adres_belgesi,
                k.adli_sicil,
                k.ehliyet,
                k.kimlik_belgesi,
                k.fotograf,
                k.basvuru_formu,
                k.e_src_kaydi,
                k.tutar,
                k.kalan,
                k.odeme_1,
                k.tarih_1,
                k.odeme_2,
                k.tarih_2,
                k.odeme_3,
                k.tarih_3,
                k.odeme_4,
                k.tarih_4,
                k.odeme_5,
                k.tarih_5,
                k.odeme_6,
                k.tarih_6
            FROM kursiyer k
            WHERE k.kurs_durumu = 'Beklemede'
            AND (
                -- Eksik evrakı olanlar
                k.ogrenim_belgesi = 0 OR
                k.adres_belgesi = 0 OR
                k.adli_sicil = 0 OR
                k.ehliyet = 0 OR
                k.kimlik_belgesi = 0 OR
                k.fotograf = 0 OR
                k.basvuru_formu = 0 OR
                k.e_src_kaydi = 0 OR
                -- VEYA ödemesi tam olmayanlar (kalan > 0)
                (k.kalan > 0 AND k.kalan IS NOT NULL)
            )
            ORDER BY k.aday_ismi
        """)
        rows = cursor.fetchall()
        result = []
        for row in rows:
            result.append({
                "id": row["id"],
                "aday_ismi": row["aday_ismi"],
                "tc_kimlik": row["tc_kimlik"],
                "tel_no": row["tel_no"],
                "sinif": row["sinif"],
                "alacagi_egitim": row["alacagi_egitim"],
                "ogrenim_belgesi": row["ogrenim_belgesi"],
                "adres_belgesi": row["adres_belgesi"],
                "adli_sicil": row["adli_sicil"],
                "ehliyet": row["ehliyet"],
                "kimlik_belgesi": row["kimlik_belgesi"],
                "fotograf": row["fotograf"],
                "basvuru_formu": row["basvuru_formu"],
                "e_src_kaydi": row["e_src_kaydi"],
                "tutar": row["tutar"],
                "kalan": row["kalan"],
                "odeme_1": row["odeme_1"],
                "tarih_1": row["tarih_1"],
                "odeme_2": row["odeme_2"],
                "tarih_2": row["tarih_2"],
                "odeme_3": row["odeme_3"],
                "tarih_3": row["tarih_3"],
                "odeme_4": row["odeme_4"],
                "tarih_4": row["tarih_4"],
                "odeme_5": row["odeme_5"],
                "tarih_5": row["tarih_5"],
                "odeme_6": row["odeme_6"],
                "tarih_6": row["tarih_6"]
            })
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Rapor alınamadı: {str(e)}")
    conn.close()
    return result

@app.delete("/api/loglar/temizle")
async def loglari_temizle(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM log")
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Loglar silinemedi: {str(e)}")
    conn.close()
    return {"detail": "Tüm loglar silindi."}

# Test endpoint
@app.get("/")
async def root():
    return {"message": "Bandırma CRM API çalışıyor"}

@app.exception_handler(404)
async def not_found_handler(request: Request, exc):
    return JSONResponse(
        status_code=404,
        content={"detail": f"Endpoint {request.url.path} not found."}
    )

@app.exception_handler(HTTPException)
async def custom_http_exception_handler(request: Request, exc: HTTPException):
    # Eğer 401 ise özel mesajı döndür
    if exc.status_code == 401:
        return JSONResponse(
            status_code=401,
            content={"detail": "Oturum süresi doldu, lütfen tekrar giriş yapınız."}
        )
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail}
    )

@app.get("/users/me/", response_model=User)
async def read_users_me(current_user: dict = Depends(get_current_user)):
    return {
        "id": current_user["id"],
        "username": current_user["kullanici_adi"],
        "role": current_user["rol"]
    }

@app.post("/logout")
async def logout(current_user: dict = Depends(get_current_user)):
    # Log logout activity for all users
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Çıkış yaptı")
        conn.commit()
        conn.close()
    
    return {"detail": "Başarıyla çıkış yapıldı."}

@app.post("/api/siniflar/yukle")
async def siniflar_excel_yukle(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    # Dosyayı geçici olarak kaydet
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    imported = 0
    errors = []
    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Sadece bu sütunlar kullanılacak
        valid_headers = ["sinif_isim", "kursiyer_sayi", "kursiyer_list"]
        header_idx = {h: i for i, h in enumerate(headers) if h in valid_headers}
        if not header_idx:
            raise Exception("Excel başlıkları uygun değil. Gerekli: sinif_isim, kursiyer_sayi, kursiyer_list")
        conn = get_db_connection()
        cursor = conn.cursor()
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                row_data = {}
                for h in valid_headers:
                    if h in header_idx:
                        val = row[header_idx[h]].value
                        row_data[h] = val if val is not None else ""
                    else:
                        row_data[h] = ""
                # En az sinif_isim dolu olmalı
                if not row_data["sinif_isim"] or str(row_data["sinif_isim"]).strip() == "":
                    continue
                # kursiyer_sayi int olarak kaydedilecek
                kursiyer_sayi_val = row_data.get("kursiyer_sayi", 0)
                try:
                    kursiyer_sayi_val = int(kursiyer_sayi_val)
                except Exception:
                    kursiyer_sayi_val = 0
                # kursiyer_list string olarak kaydedilecek
                kursiyer_list_val = row_data.get("kursiyer_list", "")
                if isinstance(kursiyer_list_val, list):
                    kursiyer_list_val = str(kursiyer_list_val)
                else:
                    kursiyer_list_val = str(kursiyer_list_val or "")
                # Ekle
                cursor.execute(
                    "INSERT INTO sinif (sinif_isim, kursiyer_sayi, kursiyer_list) VALUES (?, ?, ?)",
                    (
                        str(row_data.get("sinif_isim", "")),
                        kursiyer_sayi_val,
                        kursiyer_list_val
                    )
                )
                imported += 1
            except Exception as e:
                errors.append(f"{idx}. satır: {str(e)}")
        conn.commit()
        conn.close()
    except Exception as e:
        errors.append(f"Dosya okunamadı veya işlenemedi: {str(e)}")
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
    return {"imported": imported, "errors": errors}

@app.get("/api/belge/indir-zip/{kursiyer_id}")
async def belge_zip_indir(kursiyer_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kursiyer WHERE id = ? OR tc_kimlik = ?", (kursiyer_id, kursiyer_id))
    kursiyer = cursor.fetchone()
    if not kursiyer:
        conn.close()
        raise HTTPException(status_code=404, detail="Kursiyer bulunamadı")
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer["id"],))
    belge_row = cursor.fetchone()
    conn.close()
    if not belge_row:
        raise HTTPException(status_code=404, detail="Belge bulunamadı")

    aday_ismi = (kursiyer["aday_ismi"] or "").strip()
    alacagi_egitim = (kursiyer["alacagi_egitim"] or "").strip().upper()
    klasor_adi = f"{aday_ismi} - {alacagi_egitim}".strip(" -")
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for key, dosya_adi in [
            ("foto", "FOTOGRAF"),
            ("ogrenim_belgesi", "OGRENIM BELGESI"),
            ("adli_sicil", "ADLI SICIL"),
        ]:
            rel_path = belge_row[key] if key in belge_row.keys() else None
            if rel_path and os.path.isfile(rel_path):
                ext = os.path.splitext(rel_path)[1]
                arcname = f"{klasor_adi}/{dosya_adi}{ext}"
                zipf.write(rel_path, arcname=arcname)
    zip_buffer.seek(0)
    # --- Fix: sanitize filename for ASCII-only Content-Disposition ---
    def ascii_filename(s):
        s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
        s = re.sub(r'[^A-Za-z0-9_\-\. ]+', '', s)
        return s.strip().replace(' ', '_')
    ascii_zip_filename = ascii_filename(f"{klasor_adi}.zip")
    # Optionally, provide the original filename as filename* (RFC 5987)
    content_disposition = f'attachment; filename="{ascii_zip_filename}"'
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": content_disposition}
    )

@app.post("/api/belge/sil/{kursiyer_id}")
async def belge_sil(
    kursiyer_id: str,
    req: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    belge_tipi = req.get("belge_tipi")
    allowed_keys = [
        "foto", "ogrenim_belgesi", "adres_belgesi", "adli_sicil",
        "ehliyet", "kimlik_belgesi", "basvuru_formu"
    ]
    if belge_tipi not in allowed_keys:
        raise HTTPException(status_code=400, detail="Geçersiz belge tipi")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kursiyer WHERE id = ? OR tc_kimlik = ?", (kursiyer_id, kursiyer_id))
    kursiyer = cursor.fetchone()
    if not kursiyer:
        conn.close()
        raise HTTPException(status_code=404, detail="Kursiyer bulunamadı")
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer["id"],))
    belge_row = cursor.fetchone()
    if not belge_row:
        conn.close()
        return {}
    # Dosya yolunu bul ve sil
    rel_path = belge_row[belge_tipi] if belge_tipi in belge_row.keys() else None
    if rel_path:
        abs_path = os.path.join(os.getcwd(), rel_path)
        try:
            if os.path.isfile(abs_path):
                os.remove(abs_path)
        except Exception:
            pass
    # Veritabanında alanı null yap
    cursor.execute(
        f"UPDATE belge SET {belge_tipi} = NULL WHERE kursiyer_id = ?",
        (kursiyer["id"],)
    )
    conn.commit()
    # Güncel belge bilgisini dön
    cursor.execute("SELECT * FROM belge WHERE kursiyer_id = ?", (kursiyer["id"],))
    belge_row = cursor.fetchone()
    conn.close()
    return dict(belge_row) if belge_row else {}

@app.get("/kursiyer/excel-indir")
async def kursiyer_excel_indir(current_user: dict = Depends(get_current_user)):
    if current_user["rol"] not in ("admin", "user"):
        raise HTTPException(status_code=403, detail="Yetkisiz erişim")
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM kursiyer")
    rows = cursor.fetchall()
    
    # Log excel indirme activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Kursiyer Excel raporu indirdi")
    
    conn.close()
    if not rows:
        raise HTTPException(status_code=404, detail="Kursiyer bulunamadı")
    # Excel oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kursiyerler"
    # Sütun başlıkları
    headers = list(rows[0].keys())
    ws.append(headers)
    # Satırları ekle
    for row in rows:
        ws.append([row[h] for h in headers])
    # Dosyayı belleğe kaydet
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="kursiyerler.xlsx"'}
    )

@app.get("/api/rapor/katilanlar")
async def rapor_katilanlar(current_user: dict = Depends(get_current_user)):
    """
    kurs_durumu = 'Katiliyor' olan kursiyerler ve en son girilen sınavı (yazılı/uygulama) ile birlikte döndürür.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    
    # Log rapor alma activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Katılanlar raporu aldı")
    
    cursor.execute("SELECT * FROM kursiyer WHERE kurs_durumu = 'Katiliyor'")
    kursiyerler = cursor.fetchall()
    result = []
    for k in kursiyerler:
        kursiyer_id = k["id"]
        cursor.execute("SELECT * FROM sinav WHERE kursiyer_id = ?", (kursiyer_id,))
        sinav = cursor.fetchone()
        # En son girilen sınavı bul (tarihli yazılı/uygulama)
        last_exam = None
        last_exam_date = None
        if sinav:
            # Yazılı ve uygulama sınavlarını ve tarihlerini sırala
            exam_fields = []
            for i in range(1, 5):
                if sinav[f"sinav_{i}"] not in (None, "") and sinav[f"sinav_{i}_tarih"] not in (None, ""):
                    exam_fields.append((sinav[f"sinav_{i}_tarih"], f"Yazılı {i}", sinav[f"sinav_{i}"]))
                if sinav[f"uygulama_{i}"] not in (None, "") and sinav[f"uygulama_{i}_tarih"] not in (None, ""):
                    exam_fields.append((sinav[f"uygulama_{i}_tarih"], f"Uygulama {i}", sinav[f"uygulama_{i}"]))
            # Tarihe göre en sonu bul
            if exam_fields:
                exam_fields = [e for e in exam_fields if e[0]]
                exam_fields.sort(key=lambda x: x[0], reverse=True)
                if exam_fields:
                    last_exam_date, last_exam_type, last_exam_value = exam_fields[0]
                    last_exam = {"tarih": last_exam_date, "tip": last_exam_type, "deger": last_exam_value}
        result.append({
            "aday_ismi": k["aday_ismi"],
            "tc_kimlik": k["tc_kimlik"],
            "tel_no": k["tel_no"],
            "sinif": k["sinif"] if "sinif" in k.keys() else "",
            "last_exam": last_exam
        })
    conn.close()
    return result

@app.get("/api/rapor/farklar")
async def rapor_farklar(current_user: dict = Depends(get_current_user)):
    """
    kurs_durumu = 'Fark' olan kursiyerleri döndürür.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    
    # Log rapor alma activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Farklar raporu aldı")
    
    cursor.execute("SELECT * FROM kursiyer WHERE kurs_durumu = 'Fark'")
    rows = cursor.fetchall()
    result = []
    for row in rows:
        result.append({
            "aday_ismi": row["aday_ismi"],
            "tc_kimlik": row["tc_kimlik"],
            "tel_no": row["tel_no"],
            "sinif": row["sinif"] if "sinif" in row.keys() else "",
            "alacagi_egitim": row["alacagi_egitim"] if "alacagi_egitim" in row.keys() else "",
            "devam_egitimi": row["devam_egitimi"] if "devam_egitimi" in row.keys() else ""
        })
    conn.close()
    return result

@app.get("/api/rapor/beklemede-tam-olanlar")
async def rapor_beklemede_tam_olanlar(current_user: dict = Depends(get_current_user)):
    """
    kurs_durumu = 'Beklemede' olup evrak ve ödeme TAM olan kursiyerleri döndürür.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    
    # Log rapor alma activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Beklemede tam olanlar raporu aldı")
    
    cursor.execute("SELECT * FROM kursiyer WHERE kurs_durumu = 'Beklemede'")
    rows = cursor.fetchall()
    result = []
    for row in rows:
        # Evraklar tam mı?
        evraklar_tam = all(row[key] == 1 for key in [
            "ogrenim_belgesi", "adres_belgesi", "adli_sicil", "ehliyet", "kimlik_belgesi", "fotograf", "basvuru_formu", "e_src_kaydi"
        ])
        # Ödeme tam mı?
        odeme_tam = (row["kalan"] == 0)
        if evraklar_tam and odeme_tam:
            result.append({
                "aday_ismi": row["aday_ismi"],
                "tc_kimlik": row["tc_kimlik"],
                "tel_no": row["tel_no"],
                "sinif": row["sinif"] if "sinif" in row.keys() else "",
                "alacagi_egitim": row["alacagi_egitim"] if "alacagi_egitim" in row.keys() else "",
                "devam_egitimi": row["devam_egitimi"] if "devam_egitimi" in row.keys() else ""
            })
    conn.close()
    return result

@app.get("/api/rapor/muhasebe")
async def rapor_muhasebe(current_user: dict = Depends(get_current_user)):
    """
    Kalan borcu olan (kalan > 0) tüm kursiyerleri döndürür.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    cursor = conn.cursor()
    
    # Log rapor alma activity
    log_admin_activity(cursor, current_user["kullanici_adi"], "Sistem", "Muhasebe raporu aldı")
    
    cursor.execute("SELECT * FROM kursiyer WHERE kalan > 0 ORDER BY kalan DESC")
    rows = cursor.fetchall()
    result = []
    for row in rows:
        result.append({
            "aday_ismi": row["aday_ismi"],
            "tc_kimlik": row["tc_kimlik"],
            "tel_no": row["tel_no"],
            "sinif": row["sinif"] if "sinif" in row.keys() else "",
            "alacagi_egitim": row["alacagi_egitim"] if "alacagi_egitim" in row.keys() else "",
            "devam_egitimi": row["devam_egitimi"] if "devam_egitimi" in row.keys() else "",
            "tutar": row["tutar"] if "tutar" in row.keys() else 0,
            "kalan": row["kalan"] if "kalan" in row.keys() else 0,
            "odeme_durumu": row["odeme_durumu"] if "odeme_durumu" in row.keys() else ""
        })
    conn.close()
    return result

# --- GELEN EVRAK LOGIC (from evrak_gelen.py) ---
UPLOAD_DIR = 'uploads/evrak/gelen/'
os.makedirs(UPLOAD_DIR, exist_ok=True)

# --- GIDEN EVRAK LOGIC ---
GIDEN_UPLOAD_DIR = 'uploads/evrak/giden/'
os.makedirs(GIDEN_UPLOAD_DIR, exist_ok=True)

def get_db():
    conn = sqlite3.connect('bandirma.db')
    conn.row_factory = sqlite3.Row
    return conn

def get_evrak_with_belgeler(row):
    return {
        **dict(row),
        "belgeler": [
            {"id": b[0], "belge_yolu": b[1], "belge_adi": b[2]}
            for b in get_db().execute(
                "SELECT id, belge_yolu, belge_adi FROM gelen_evrak_belge WHERE evrak_id = ?", (row["id"],)
            ).fetchall()
        ]
    }

def get_giden_evrak_with_belgeler(row):
    evrak_eki = row["evrak_eki"] if "evrak_eki" in row.keys() else 0
    belgeler = [
        {"id": b[0], "belge_yolu": b[1], "belge_adi": b[2]}
        for b in get_db().execute(
            "SELECT id, belge_yolu, belge_adi FROM giden_evrak_belge WHERE evrak_id = ? ORDER BY id ASC",
            (row["id"],)
        ).fetchall()
    ]
    # Slot sayısı kadar doldur
    while len(belgeler) < evrak_eki:
        belgeler.append({})
    return {
        **dict(row),
        "belgeler": belgeler[:evrak_eki]  # fazlaysa kırp
    }

def get_belge_list(belgeler: Union[UploadFile, List[UploadFile], None] = File(None)):
    if belgeler is None:
        return []
    if isinstance(belgeler, list):
        return belgeler
    return [belgeler]

@app.get("/gelen-evrak", response_model=List[dict])
def list_gelen_evrak():
    db = get_db()
    rows = db.execute("SELECT * FROM gelen_evrak ORDER BY sira_no ASC").fetchall()
    return [get_evrak_with_belgeler(row) for row in rows]

@app.get("/gelen-evrak/{evrak_id}", response_model=dict)
def get_gelen_evrak(evrak_id: int):
    db = get_db()
    row = db.execute("SELECT * FROM gelen_evrak WHERE id = ?", (evrak_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Evrak bulunamadı")
    return get_evrak_with_belgeler(row)

@app.post("/gelen-evrak")
def create_gelen_evrak(
    sira_no: int = Form(...),
    kimden: str = Form(...),
    evrak_tarih: str = Form(...),
    evrak_no: str = Form(...),
    evrak_cinsi: str = Form(...),
    evrak_eki: int = Form(...),
    evrak_ozet: str = Form(...),
    belgeler: Optional[List[UploadFile]] = File(None),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Sadece admin kullanıcı evrak ekleyebilir.")
    try:
        print("[GELEN EVRAK KAYIT] Form Data:", {
            'sira_no': sira_no,
            'kimden': kimden,
            'evrak_tarih': evrak_tarih,
            'evrak_no': evrak_no,
            'evrak_cinsi': evrak_cinsi,
            'evrak_eki': evrak_eki,
            'evrak_ozet': evrak_ozet,
            'belge_count': len(belgeler) if belgeler else 0
        })
        db = get_db()
        cur = db.cursor()
        cur.execute(
            """
            INSERT INTO gelen_evrak (sira_no, kimden, evrak_tarih, evrak_no, evrak_cinsi, evrak_eki, evrak_ozet)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (sira_no, kimden, evrak_tarih, evrak_no, evrak_cinsi, evrak_eki, evrak_ozet)
        )
        evrak_id = cur.lastrowid
        belge_list = []
        belge_count = len(belgeler) if belgeler else 0
        for i in range(evrak_eki):
            if belgeler and i < belge_count and belgeler[i] and getattr(belgeler[i], 'filename', None):
                belge = belgeler[i]
                filename = getattr(belge, "filename", "")
                if not filename or not isinstance(filename, str):
                    cur.execute(
                        "INSERT INTO gelen_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, NULL, NULL)",
                        (evrak_id,)
                    )
                    continue
                filename = str(filename)
                if not filename:
                    cur.execute(
                        "INSERT INTO gelen_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, NULL, NULL)",
                        (evrak_id,)
                    )
                    continue
                print(f"[GELEN EVRAK KAYIT] Dosya: {filename}, ContentType: {belge.content_type}")
                # --- YENİ DOSYALAMA ---
                klasor_yolu = os.path.join("uploads", "EVRAK", "GELEN", str(sira_no))
                os.makedirs(klasor_yolu, exist_ok=True)
                file_path = os.path.join(klasor_yolu, filename)
                with open(file_path, "wb") as buffer:
                    shutil.copyfileobj(belge.file, buffer)
                cur.execute(
                    "INSERT INTO gelen_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, ?, ?)",
                    (evrak_id, file_path, filename)
                )
                belge_list.append({"belge_yolu": file_path, "belge_adi": filename})
            else:
                cur.execute(
                    "INSERT INTO gelen_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, NULL, NULL)",
                    (evrak_id,)
                )
        db.commit()
        print(f"[GELEN EVRAK KAYIT] Başarılı! Evrak ID: {evrak_id}")
        return {"id": evrak_id, "belgeler": belge_list}
    except Exception as e:
        print("[GELEN EVRAK KAYIT] HATA:", str(e))
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Kayıt sırasında sunucu hatası: {str(e)}")

@app.put("/gelen-evrak/{evrak_id}")
def update_gelen_evrak(
    evrak_id: int,
    sira_no: int = Form(...),
    kimden: str = Form(...),
    evrak_tarih: str = Form(...),
    evrak_no: str = Form(...),
    evrak_cinsi: str = Form(...),
    evrak_eki: int = Form(...),
    evrak_ozet: str = Form(...),
    belgeler: Optional[List[UploadFile]] = File(None),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Sadece admin kullanıcı evrak düzenleyebilir.")
    db = get_db()
    cur = db.cursor()
    cur.execute(
        """
        UPDATE gelen_evrak SET sira_no=?, kimden=?, evrak_tarih=?, evrak_no=?, evrak_cinsi=?, evrak_eki=?, evrak_ozet=? WHERE id=?
        """,
        (sira_no, kimden, evrak_tarih, evrak_no, evrak_cinsi, evrak_eki, evrak_ozet, evrak_id)
    )
    # --- EKLENDİ: Eksik belge slotlarını oluştur ---
    cur.execute("SELECT COUNT(*) FROM gelen_evrak_belge WHERE evrak_id=?", (evrak_id,))
    mevcut_slot = cur.fetchone()[0]
    for _ in range(mevcut_slot, evrak_eki):
        cur.execute("INSERT INTO gelen_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, NULL, NULL)", (evrak_id,))
    # --- SON ---
    belge_list = []
    if belgeler:
        for belge in belgeler:
            filename = getattr(belge, "filename", "")
            if not filename or not isinstance(filename, str):
                continue
            filename = str(filename)
            if not filename:
                continue
            print(f"[GELEN EVRAK KAYIT] Dosya: {filename}, ContentType: {belge.content_type}")
            klasor_yolu = os.path.join("uploads", "evrak", "gelen", str(sira_no))
            os.makedirs(klasor_yolu, exist_ok=True)
            file_path = os.path.join(klasor_yolu, filename)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(belge.file, buffer)
            cur.execute(
                "INSERT INTO gelen_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, ?, ?)",
                (evrak_id, file_path, filename)
            )
            belge_list.append({"belge_yolu": file_path, "belge_adi": filename})
    db.commit()
    return {"id": evrak_id, "belgeler": belge_list}

@app.delete("/gelen-evrak/{evrak_id}/belge/{belge_id}")
def delete_belge(evrak_id: int, belge_id: int):
    db = get_db()
    cur = db.cursor()
    belge = cur.execute("SELECT belge_yolu FROM gelen_evrak_belge WHERE id=? AND evrak_id=?", (belge_id, evrak_id)).fetchone()
    if not belge:
        raise HTTPException(status_code=404, detail="Belge bulunamadı")
    file_path = belge[0]
    if os.path.exists(file_path):
        os.remove(file_path)
    cur.execute("DELETE FROM gelen_evrak_belge WHERE id=? AND evrak_id=?", (belge_id, evrak_id))
    db.commit()
    return {"ok": True}

# --- GIDEN EVRAK ENDPOINTS ---

@app.get("/giden-evrak", response_model=List[dict])
def list_giden_evrak():
    db = get_db()
    rows = db.execute("SELECT * FROM giden_evrak ORDER BY sira_no ASC").fetchall()
    return [get_giden_evrak_with_belgeler(row) for row in rows]

@app.get("/giden-evrak/{evrak_id}", response_model=dict)
def get_giden_evrak(evrak_id: int):
    db = get_db()
    row = db.execute("SELECT * FROM giden_evrak WHERE id = ?", (evrak_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Evrak bulunamadı")
    return get_giden_evrak_with_belgeler(row)

@app.post("/giden-evrak")
def create_giden_evrak(
    sira_no: int = Form(...),
    nereye: str = Form(...),
    evrak_tarih: str = Form(...),
    evrak_no: str = Form(...),
    evrak_cinsi: str = Form(...),
    evrak_eki: int = Form(...),
    evrak_ozet: str = Form(...),
    belgeler: List[UploadFile] = Depends(get_belge_list),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Sadece admin kullanıcı evrak ekleyebilir.")
    try:
        print("[GIDEN EVRAK KAYIT] Form Data:", {
            'sira_no': sira_no,
            'nereye': nereye,
            'evrak_tarih': evrak_tarih,
            'evrak_no': evrak_no,
            'evrak_cinsi': evrak_cinsi,
            'evrak_eki': evrak_eki,
            'evrak_ozet': evrak_ozet,
            'belge_count': len(belgeler)
        })
        print("[GIDEN EVRAK KAYIT] Belgeler:", belgeler)
        print("[GIDEN EVRAK KAYIT] Belgeler type:", type(belgeler))
        if belgeler:
            for i, belge in enumerate(belgeler):
                print(f"[GIDEN EVRAK KAYIT] Belge {i}:", {
                    'filename': getattr(belge, 'filename', 'None'),
                    'content_type': getattr(belge, 'content_type', 'None'),
                    'size': getattr(belge, 'size', 'None'),
                    'type': type(belge)
                })
        db = get_db()
        cur = db.cursor()
        cur.execute(
            """
            INSERT INTO giden_evrak (sira_no, nereye, evrak_tarih, evrak_no, evrak_cinsi, evrak_eki, evrak_ozet)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (sira_no, nereye, evrak_tarih, evrak_no, evrak_cinsi, evrak_eki, evrak_ozet)
        )
        evrak_id = cur.lastrowid
        
        # Evrak_eki kadar slot oluştur
        for i in range(evrak_eki):
            cur.execute("INSERT INTO giden_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, NULL, NULL)", (evrak_id,))
        
        # Gelen dosyaları sırayla slotlara yerleştir
        belge_idx = 0
        cur.execute("SELECT id FROM giden_evrak_belge WHERE evrak_id=? ORDER BY id ASC", (evrak_id,))
        slotlar = cur.fetchall()
        print(f"[GIDEN EVRAK KAYIT] Oluşturulan slot sayısı: {len(slotlar)}")
        
        for i, slot in enumerate(slotlar):
            print(f"[GIDEN EVRAK KAYIT] Slot {i} işleniyor, belge_idx: {belge_idx}")
            if belgeler and belge_idx < len(belgeler) and belgeler[belge_idx]:
                belge = belgeler[belge_idx]
                print(f"[GIDEN EVRAK KAYIT] Belge bulundu: {getattr(belge, 'filename', 'None')}")
                
                try:
                    belge.file.seek(0, 2)  # EOF'a git
                    size = belge.file.tell()
                    belge.file.seek(0)
                    print(f"[GIDEN EVRAK KAYIT] Dosya boyutu: {size}")
                    
                    if size == 0:
                        print(f"[GIDEN EVRAK KAYIT] Dosya boyutu 0, atlanıyor")
                        belge_idx += 1
                        continue
                        
                    filename = str(getattr(belge, "filename", ""))
                    if not filename:
                        print(f"[GIDEN EVRAK KAYIT] Dosya adı boş, atlanıyor")
                        belge_idx += 1
                        continue
                        
                    print(f"[GIDEN EVRAK KAYIT] Dosya işleniyor: {filename}, ContentType: {belge.content_type}")
                    klasor_yolu = os.path.join("uploads", "evrak", "giden", str(sira_no))
                    os.makedirs(klasor_yolu, exist_ok=True)
                    file_path = os.path.join(klasor_yolu, filename)
                    print(f"[GIDEN EVRAK KAYIT] Dosya yolu: {file_path}")
                    
                    with open(file_path, "wb") as buffer:
                        shutil.copyfileobj(belge.file, buffer)
                    
                    cur.execute(
                        "UPDATE giden_evrak_belge SET belge_yolu=?, belge_adi=? WHERE id=?",
                        (file_path, filename, slot[0])
                    )
                    print(f"[GIDEN EVRAK KAYIT] Dosya başarıyla kaydedildi: {file_path}")
                    belge_idx += 1
                    
                except Exception as e:
                    print(f"[GIDEN EVRAK KAYIT] Dosya işleme hatası: {str(e)}")
                    belge_idx += 1
                    continue
            else:
                print(f"[GIDEN EVRAK KAYIT] Slot {i} için belge yok veya boş")
                belge_idx += 1
        
        db.commit()
        print(f"[GIDEN EVRAK KAYIT] Başarılı! Evrak ID: {evrak_id}")
        
        # Güncel belge listesini döndür
        cur.execute("SELECT id, belge_yolu, belge_adi FROM giden_evrak_belge WHERE evrak_id=? ORDER BY id ASC", (evrak_id,))
        belge_list = [
            {"id": b[0], "belge_yolu": b[1], "belge_adi": b[2]} for b in cur.fetchall()
        ]
        return {"id": evrak_id, "belgeler": belge_list}
    except Exception as e:
        print("[GIDEN EVRAK KAYIT] HATA:", str(e))
        print("[GIDEN EVRAK KAYIT] Hata türü:", type(e).__name__)
        import traceback
        print("[GIDEN EVRAK KAYIT] Tam hata detayı:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Kayıt sırasında sunucu hatası: {str(e)}")

@app.put("/giden-evrak/{evrak_id}")
def update_giden_evrak(
    evrak_id: int,
    sira_no: int = Form(...),
    nereye: str = Form(...),
    evrak_tarih: str = Form(...),
    evrak_no: str = Form(...),
    evrak_cinsi: str = Form(...),
    evrak_eki: int = Form(...),
    evrak_ozet: str = Form(...),
    belgeler: List[UploadFile] = Depends(get_belge_list),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Sadece admin kullanıcı evrak düzenleyebilir.")
    yeni_belge_count = len(belgeler)
    db = get_db()
    cur = db.cursor()
    # Evrak ana verisini güncelle
    cur.execute(
        """
        UPDATE giden_evrak SET sira_no=?, nereye=?, evrak_tarih=?, evrak_no=?, evrak_cinsi=?, evrak_eki=?, evrak_ozet=? WHERE id=?
        """,
        (sira_no, nereye, evrak_tarih, evrak_no, evrak_cinsi, evrak_eki, evrak_ozet, evrak_id)
    )
    # --- EKLENDİ: Eksik belge slotlarını oluştur ---
    cur.execute("SELECT COUNT(*) FROM giden_evrak_belge WHERE evrak_id=?", (evrak_id,))
    mevcut_slot = cur.fetchone()[0]
    for _ in range(mevcut_slot, evrak_eki):
        cur.execute("INSERT INTO giden_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, NULL, NULL)", (evrak_id,))
    # --- SON ---
    # Slotları sırayla yönet
    # Önce mevcut belgeleri çek
    cur.execute("SELECT id, belge_yolu, belge_adi FROM giden_evrak_belge WHERE evrak_id=? ORDER BY id ASC", (evrak_id,))
    mevcut_belgeler = cur.fetchall()
    mevcut_belge_count = len(mevcut_belgeler)
    # Evrak_eki kadar slotu koru
    # 1. Slotlar eksikse NULL slot ekle
    for i in range(mevcut_belge_count, evrak_eki):
        cur.execute("INSERT INTO giden_evrak_belge (evrak_id, belge_yolu, belge_adi) VALUES (?, NULL, NULL)", (evrak_id,))
    # 2. Slot fazlaysa fazla slotları sil
    if mevcut_belge_count > evrak_eki:
        cur.execute("SELECT id, belge_yolu FROM giden_evrak_belge WHERE evrak_id=? ORDER BY id ASC", (evrak_id,))
        all_belgeler = cur.fetchall()
        for b in all_belgeler[evrak_eki:]:
            if b[1] and os.path.exists(b[1]):
                os.remove(b[1])
            cur.execute("DELETE FROM giden_evrak_belge WHERE id=?", (b[0],))
    # 3. Güncel slotları tekrar çek
    cur.execute("SELECT id, belge_yolu, belge_adi FROM giden_evrak_belge WHERE evrak_id=? ORDER BY id ASC", (evrak_id,))
    slotlar = cur.fetchall()
    # 4. Gelen dosyaları sırayla slotlara yerleştir
    belge_idx = 0
    for i, slot in enumerate(slotlar):
        if belgeler and belge_idx < len(belgeler) and belgeler[belge_idx]:
            belge = belgeler[belge_idx]
            belge.file.seek(0, 2)  # EOF'a git
            size = belge.file.tell()
            belge.file.seek(0)
            if size == 0:
                cur.execute(
                    "UPDATE giden_evrak_belge SET belge_yolu=NULL, belge_adi=NULL WHERE id=?",
                    (slot[0],)
                )
                belge_idx += 1
                continue
            filename = str(getattr(belge, "filename", ""))
            if not filename:
                belge_idx += 1
                continue
            if slot[1] and os.path.exists(slot[1]):
                os.remove(slot[1])
            klasor_yolu = os.path.join("uploads", "evrak", "giden", str(sira_no))
            os.makedirs(klasor_yolu, exist_ok=True)
            file_path = os.path.join(klasor_yolu, filename)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(belge.file, buffer)
            cur.execute(
                "UPDATE giden_evrak_belge SET belge_yolu=?, belge_adi=? WHERE id=?",
                (file_path, filename, slot[0])
            )
            belge_idx += 1
        else:
            cur.execute(
                "UPDATE giden_evrak_belge SET belge_yolu=NULL, belge_adi=NULL WHERE id=?",
                (slot[0],)
            )
            belge_idx += 1
    db.commit()
    cur.execute("SELECT id, belge_yolu, belge_adi FROM giden_evrak_belge WHERE evrak_id=? ORDER BY id ASC", (evrak_id,))
    belge_list = [
        {"id": b[0], "belge_yolu": b[1], "belge_adi": b[2]} for b in cur.fetchall()
    ]
    return {"id": evrak_id, "belgeler": belge_list}

@app.delete("/giden-evrak/{evrak_id}/belge/{belge_id}")
def delete_giden_belge(evrak_id: int, belge_id: int):
    db = get_db()
    cur = db.cursor()
    # Önce dosya yolunu bul
    cur.execute("SELECT belge_yolu FROM giden_evrak_belge WHERE id=? AND evrak_id=?", (belge_id, evrak_id))
    belge = cur.fetchone()
    if belge and belge[0]:
        try:
            if os.path.exists(belge[0]):
                os.remove(belge[0])
        except Exception:
            pass
    # Sadece slotu NULL yap
    cur.execute("UPDATE giden_evrak_belge SET belge_yolu=NULL, belge_adi=NULL WHERE id=? AND evrak_id=?", (belge_id, evrak_id))
    # Ardından, bu evraka ait tüm NULL slotları sil
    cur.execute("DELETE FROM giden_evrak_belge WHERE evrak_id=? AND belge_yolu IS NULL AND belge_adi IS NULL", (evrak_id,))
    db.commit()
    return {"ok": True}

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    import sys
    print("422 Validation Error:", exc.errors(), file=sys.stderr)
    # print("Body:", await request.body(), file=sys.stderr)  # KALDIRILDI
    return await request_validation_exception_handler(request, exc)

@app.put("/gelen-evrak/{evrak_id}/belge/{belge_id}")
def update_gelen_evrak_belge(
    evrak_id: int,
    belge_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Sadece admin kullanıcı belge güncelleyebilir.")
    db = get_db()
    cur = db.cursor()
    # Evrak ve sıra_no bul
    cur.execute("SELECT sira_no FROM gelen_evrak WHERE id=?", (evrak_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Evrak bulunamadı")
    sira_no = row[0]
    # Dosya kaydet (benzersiz isim)
    klasor_yolu = os.path.join("uploads", "evrak", "gelen", str(sira_no))
    os.makedirs(klasor_yolu, exist_ok=True)
    filename = str(getattr(file, "filename", ""))
    base, ext = os.path.splitext(filename)
    file_path = os.path.join(klasor_yolu, filename)
    counter = 2
    while os.path.exists(file_path):
        filename = f"{base}{counter}{ext}"
        file_path = os.path.join(klasor_yolu, filename)
        counter += 1
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    # Veritabanında güncelle
    cur.execute(
        "UPDATE gelen_evrak_belge SET belge_yolu=?, belge_adi=? WHERE id=? AND evrak_id=?",
        (file_path, filename, belge_id, evrak_id)
    )
    db.commit()
    return {"ok": True, "belge_yolu": file_path, "belge_adi": filename}

@app.put("/giden-evrak/{evrak_id}/belge/{belge_id}")
def update_giden_evrak_belge(
    evrak_id: int,
    belge_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Sadece admin kullanıcı belge güncelleyebilir.")
    db = get_db()
    cur = db.cursor()
    # Evrak ve sıra_no bul
    cur.execute("SELECT sira_no FROM giden_evrak WHERE id=?", (evrak_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Evrak bulunamadı")
    sira_no = row[0]
    # Dosya kaydet (benzersiz isim)
    klasor_yolu = os.path.join("uploads", "EVRAK", "GIDEN", str(sira_no))
    os.makedirs(klasor_yolu, exist_ok=True)
    filename = str(getattr(file, "filename", ""))
    base, ext = os.path.splitext(filename)
    file_path = os.path.join(klasor_yolu, filename)
    counter = 2
    while os.path.exists(file_path):
        filename = f"{base}{counter}{ext}"
        file_path = os.path.join(klasor_yolu, filename)
        counter += 1
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    # Veritabanında güncelle
    cur.execute(
        "UPDATE giden_evrak_belge SET belge_yolu=?, belge_adi=? WHERE id=? AND evrak_id=?",
        (file_path, filename, belge_id, evrak_id)
    )
    db.commit()
    return {"ok": True, "belge_yolu": file_path, "belge_adi": filename}

@app.delete("/gelen-evrak/{evrak_id}")
def delete_gelen_evrak(
    evrak_id: int,
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Sadece admin kullanıcı evrak silebilir.")
    db = get_db()
    cur = db.cursor()
    # Evrakı bul
    cur.execute("SELECT sira_no FROM gelen_evrak WHERE id=?", (evrak_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Evrak bulunamadı")
    sira_no = row[0]
    # Belgeleri bul ve sil
    cur.execute("SELECT belge_yolu FROM gelen_evrak_belge WHERE evrak_id=?", (evrak_id,))
    belgeler = cur.fetchall()
    for belge in belgeler:
        if belge[0] and os.path.exists(belge[0]):
            try:
                os.remove(belge[0])
            except Exception:
                pass
    # Belgeler klasörünü sil
    klasor_yolu = os.path.join("uploads", "evrak", "gelen", str(sira_no))
    if os.path.exists(klasor_yolu):
        import shutil
        try:
            shutil.rmtree(klasor_yolu)
        except Exception:
            pass
    # Veritabanından belge kayıtlarını ve evrağı sil
    cur.execute("DELETE FROM gelen_evrak_belge WHERE evrak_id=?", (evrak_id,))
    cur.execute("DELETE FROM gelen_evrak WHERE id=?", (evrak_id,))
    db.commit()
    return {"ok": True}

@app.delete("/giden-evrak/{evrak_id}")
def delete_giden_evrak(
    evrak_id: int,
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Sadece admin kullanıcı evrak silebilir.")
    db = get_db()
    cur = db.cursor()
    # Evrakı bul
    cur.execute("SELECT sira_no FROM giden_evrak WHERE id=?", (evrak_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Evrak bulunamadı")
    sira_no = row[0]
    # Belgeleri bul ve sil
    cur.execute("SELECT belge_yolu FROM giden_evrak_belge WHERE evrak_id=?", (evrak_id,))
    belgeler = cur.fetchall()
    for belge in belgeler:
        if belge[0] and os.path.exists(belge[0]):
            try:
                os.remove(belge[0])
            except Exception:
                pass
    # Belgeler klasörünü sil
    klasor_yolu = os.path.join("uploads", "evrak", "giden", str(sira_no))
    if os.path.exists(klasor_yolu):
        import shutil
        try:
            shutil.rmtree(klasor_yolu)
        except Exception:
            pass
    # Veritabanından belge kayıtlarını ve evrağı sil
    cur.execute("DELETE FROM giden_evrak_belge WHERE evrak_id=?", (evrak_id,))
    cur.execute("DELETE FROM giden_evrak WHERE id=?", (evrak_id,))
    db.commit()
    return {"ok": True}

@app.patch("/kursiyer/{kid}")
async def kursiyer_patch(kid: str, req: dict, current_user: dict = Depends(get_current_user)):
    """
    Kursiyer bilgilerini kısmi olarak günceller (PATCH endpoint)
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Veritabanı bağlantı hatası")
    
    cursor = conn.cursor()
    
    # Kursiyer ID'sini kontrol et
    field = "id"
    cursor.execute("PRAGMA table_info(kursiyer)")
    columns = [col[1] for col in cursor.fetchall()]
    if "id" not in columns:
        field = "tc_kimlik"
    
    # Kursiyer var mı kontrol et
    cursor.execute(f"SELECT * FROM kursiyer WHERE {field} = ?", (kid,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Kursiyer bulunamadı")
    
    # Güncellenecek alanları hazırla
    update_fields = []
    values = []
    
    # Sadece gönderilen alanları güncelle
    for key, value in req.items():
        if key in columns and key != "id":
            update_fields.append(f"{key} = ?")
            values.append(value)
    
    if not update_fields:
        conn.close()
        return {"detail": "Güncellenecek alan bulunamadı"}
    
    # Güncelleme işlemini yap
    values.append(kid)
    try:
        cursor.execute(
            f"UPDATE kursiyer SET {', '.join(update_fields)} WHERE {field} = ?",
            values
        )
        
        # Log kaydı ekle
        aday_ismi = row["aday_ismi"]
        log_admin_activity(cursor, current_user["kullanici_adi"], aday_ismi, "kursiyer güncellendi (PATCH)")
        
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Güncelleme hatası: {str(e)}")
    
    conn.close()
    return {"detail": "Kursiyer başarıyla güncellendi."}